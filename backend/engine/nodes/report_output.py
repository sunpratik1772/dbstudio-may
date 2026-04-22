import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..context import RunContext
from ..node_spec import NodeSpec, _spec
from ..ports import ParamSpec, ParamType, PortSpec, PortType


def _hex_fill(hex_colour: str) -> PatternFill:
    colour = hex_colour.lstrip("#").upper().zfill(6)
    return PatternFill(fill_type="solid", fgColor=colour)


def _write_df(ws, df: pd.DataFrame, freeze: bool = True) -> None:
    header_fill = _hex_fill("1C2333")
    header_font = Font(bold=True, color="F9FAFB", size=10)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    if freeze:
        ws.freeze_panes = "A2"

    colour_col_idx = (
        df.columns.get_loc("_highlight_colour") + 1
        if "_highlight_colour" in df.columns
        else None
    )

    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        row_colour: str | None = None
        if colour_col_idx is not None:
            c = row_vals[colour_col_idx - 1]
            if c and c not in ("#FFFFFF", ""):
                row_colour = c

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if row_colour:
                cell.fill = _hex_fill(row_colour)
                cell.font = Font(size=9, color="000000")
            else:
                cell.font = Font(size=9)

    # Auto-width (capped at 40)
    for ci, col in enumerate(df.columns, 1):
        max_len = len(str(col))
        if len(df) > 0:
            max_len = max(max_len, df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(int(max_len) + 2, 40)


def _df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Stringify datetime columns; Excel can't hold timezone-aware datetimes."""
    df = df.copy()
    for col in df.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]", "datetimetz"]).columns:
        df[col] = df[col].astype(str)
    # Lists/dicts in cells break Excel
    for col in df.columns:
        if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
            df[col] = df[col].apply(str)
    return df


def handle_report_output(node: dict, ctx: RunContext) -> None:
    cfg = node.get("config", {})
    raw_output_path: str = cfg.get("output_path", "output/report.xlsx")
    # Resolve {context.xxx} and strip any remaining {unresolved} placeholders so filenames are safe.
    output_path = ctx.inject_template(raw_output_path)
    output_path = re.sub(r"\{[^}]*\}", "", output_path)

    # On Cloud Run the repo-relative "output/" directory is not
    # writable — the container filesystem is read-only except for
    # `/tmp`. We honour `DBSHERPA_OUTPUT_DIR` so deployments can
    # redirect report writes to `/tmp/output` (ephemeral demo) or a
    # GCS FUSE mount (persistent) without touching workflow YAML.
    output_root = os.environ.get("DBSHERPA_OUTPUT_DIR")
    if output_root and not os.path.isabs(output_path):
        # Normalise by stripping a leading "output/" so the env var
        # is authoritative (avoids `/tmp/output/output/report.xlsx`).
        norm = output_path
        if norm.startswith("output/"):
            norm = norm[len("output/"):]
        output_path = os.path.join(output_root, norm)
    tabs: list[dict] = cfg.get("tabs", [])
    # "first" (default) keeps Cover → Exec Summary → Section Summaries → data tabs.
    # "last" emits data tabs first, then summary sheets at the end.
    summary_position: str = cfg.get("summary_position", "first")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def build_cover() -> None:
        ws_cover = wb.create_sheet("Cover")
        ws_cover.sheet_view.showGridLines = False
        bg = _hex_fill("0D1B2A")
        for row in ws_cover.iter_rows(min_row=1, max_row=40, min_col=1, max_col=15):
            for cell in row:
                cell.fill = bg

        def cover_cell(addr: str, text: str, bold: bool = False, size: int = 12, color: str = "F9FAFB"):
            c = ws_cover[addr]
            c.value = text
            c.font = Font(bold=bold, size=size, color=color)

        cover_cell("B2", "dbSherpa — Trade Surveillance Report", bold=True, size=22, color="F59E0B")
        cover_cell("B4", f"Trader:       {ctx.get('trader_id', 'N/A')}", size=13)
        cover_cell("B5", f"Instrument:   {ctx.get('currency_pair', 'N/A')}", size=13)
        cover_cell("B6", f"Alert Date:   {ctx.get('alert_date', 'N/A')}", size=13)
        disp_colour = {"ESCALATE": "EF4444", "REVIEW": "F59E0B", "DISMISS": "10B981"}.get(ctx.disposition, "F9FAFB")
        cover_cell("B8", f"Disposition:  {ctx.disposition}", bold=True, size=16, color=disp_colour)
        cover_cell("B9", f"Signal Flags: {ctx.get('flag_count', 0)}", size=13)
        ws_cover.column_dimensions["B"].width = 60

    def build_exec_summary() -> None:
        if not ctx.executive_summary:
            return
        ws_exec = wb.create_sheet("Executive Summary")
        ws_exec["A1"] = "EXECUTIVE SUMMARY"
        ws_exec["A1"].font = Font(bold=True, size=14, color="0D1B2A")
        ws_exec["A3"] = ctx.executive_summary
        ws_exec["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_exec.column_dimensions["A"].width = 120
        ws_exec.row_dimensions[3].height = 500

    def build_section_summaries() -> None:
        if not ctx.sections:
            return
        ws_sec = wb.create_sheet("Section Summaries")
        row = 1
        for name, sec in ctx.sections.items():
            ws_sec.cell(row=row, column=1, value=name.upper().replace("_", " ")).font = Font(bold=True, size=12)
            row += 1
            cell = ws_sec.cell(row=row, column=1, value=sec.get("narrative", ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws_sec.row_dimensions[row].height = 200
            row += 3
        ws_sec.column_dimensions["A"].width = 110

    def build_summary_sheets() -> None:
        build_cover()
        build_exec_summary()
        build_section_summaries()

    def build_data_tabs() -> None:
        local_tabs = tabs or [
            {"name": ds_name[:31], "dataset": ds_name, "include_highlights": True}
            for ds_name in ctx.datasets
            if not ds_name.endswith("_highlighted")
        ]
        for tab in local_tabs:
            tab_name: str = tab.get("name", "Data")[:31]
            ds_name: str = tab.get("dataset", "")
            use_highlights: bool = tab.get("include_highlights", True)

            df = ctx.datasets.get(ds_name)
            if df is None:
                continue

            highlighted_key = f"{ds_name}_highlighted"
            if use_highlights and highlighted_key in ctx.datasets:
                df = ctx.datasets[highlighted_key]

            ws = wb.create_sheet(tab_name)
            _write_df(ws, _df_for_excel(df))

    if summary_position == "last":
        build_data_tabs()
        build_summary_sheets()
    else:
        build_summary_sheets()
        build_data_tabs()

    wb.save(output_path)
    ctx.report_path = output_path
    ctx.set("report_path", output_path)


NODE_SPEC: NodeSpec = _spec(
    "REPORT_OUTPUT",
    handle_report_output,
    "Generate Excel report with tabs & highlights",
    color="#047857",
    icon="FileSpreadsheet",
    config_tags=("output_name",),
    input_ports=(
        PortSpec(
            name="datasets",
            type=PortType.OBJECT,
            description="All DataFrames to include as tabs (ctx.datasets).",
        ),
        PortSpec(
            name="sections",
            type=PortType.OBJECT,
            description="Section narratives for the Section Summaries sheet.",
            optional=True,
        ),
        PortSpec(
            name="executive_summary",
            type=PortType.TEXT,
            description="Executive summary text.",
            optional=True,
        ),
        PortSpec(
            name="context",
            type=PortType.OBJECT,
            description="disposition, trader_id, currency_pair etc. used on the cover page.",
            optional=True,
        ),
    ),
    output_ports=(
        PortSpec(
            name="report_path",
            type=PortType.TEXT,
            description="Absolute path to the written .xlsx file. Stored as context.report_path.",
        ),
    ),
    params=(
        ParamSpec(
            name="output_path",
            type=ParamType.STRING,
            description="File path for the Excel output (e.g. 'output/report.xlsx').",
            required=True,
        ),
        ParamSpec(
            name="tabs",
            type=ParamType.ARRAY,
            description=(
                "Array of {name: string (max 31 chars), dataset: string, "
                "include_highlights: boolean}. When empty, all context.datasets are included."
            ),
            default=[],
            required=False,
        ),
    ),
    constraints=(
        "Tab names truncated to 31 characters (Excel limit).",
        "Datetime columns converted to strings automatically.",
        "List/dict cell values stringified automatically.",
        "If include_highlights=true, uses dataset_name + '_highlighted' if it exists.",
        "Must be the final node in the workflow.",
    ),
)
